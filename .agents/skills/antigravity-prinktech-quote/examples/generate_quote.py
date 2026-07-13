import openpyxl
import os
import shutil
import win32com.client
import sys
import argparse

def create_quote(args):
    template_file = r"D:\16. Code\32-website-prinktech\.agents\skills\antigravity-prinktech-quote\resources\template_bao_gia.xlsx"
    name_slug = args.name.replace(' ', '_')
    excel_output = os.path.join(args.output_dir, f"1. Bao_gia_in_tem_UV_DTF_{name_slug}_V2.xlsx")
    pdf_output = os.path.join(args.output_dir, f"2. Bao_gia_in_tem_UV_DTF_{name_slug}_V2.pdf")

    if not os.path.exists(template_file):
        print(f"[ERR] Template not found: {template_file}")
        sys.exit(1)

    os.makedirs(args.output_dir, exist_ok=True)
    
    # 1. Sao chép template Excel
    print("Copying template file...")
    shutil.copyfile(template_file, excel_output)

    # 2. Điền thông tin vào Excel
    print("Filling customer data...")
    wb = openpyxl.load_workbook(excel_output)
    ws = wb.active

    # Cập nhật số đơn hàng ở D2
    ws["D2"] = f"Đơn hàng: {args.order_code}"
    # Khách hàng & Địa chỉ
    ws["A6"] = f" Khách hàng: {args.name}"
    ws["A7"] = f" Địa chỉ giao hàng: {args.address}"

    # Kiểm tra xem có sử dụng mét dài không
    use_meters = args.meters is not None and args.meters > 0

    if not use_meters:
        print("Order is based on pieces/sheets (no meters). Removing Column E...")
        # Unmerge toàn bộ sheet trước khi xóa cột để tránh lỗi openpyxl
        merged_ranges = list(ws.merged_cells.ranges)
        for r in merged_ranges:
            ws.unmerge_cells(r.coord)
            
        # Xoá cột E (Mét dài)
        ws.delete_cols(5, amount=1)
        
        # Cập nhật Quy cách dòng 6
        ws["D6"] = f" Quy cách: {args.spec_text or 'Tem UV DTF bọc màng định hình bế sẵn'}"

        # Điền sản phẩm (Dòng 10)
        ws["B10"] = args.product
        ws["C10"] = args.size
        ws["D10"] = args.qty
        ws["E10"] = args.rate_excl_vat # Đơn giá cái/tờ trước VAT

        # Ghi đè lại công thức dòng 10 (Số lượng × Đơn giá cái)
        ws["F10"] = "=D10*E10" # Thành tiền chưa VAT
        ws["G10"] = "=F10/D10" # Đơn giá cái chưa VAT

        # Ghi đè các công thức tính tổng dòng 11
        ws["D11"] = "=SUM(D10:D10)"
        ws["F11"] = "=SUM(F10:F10)"
        ws["G11"] = "=F11/D11"

        # Ghi đè công thức khối VAT (dòng 13-17) ở cột G (cột 7 mới)
        ws["G13"] = "=F11" # Cộng tiền hàng
        ws["G14"] = "=G13*0.08" # Thuế VAT
        ws["G15"] = "=G13+G14" # Tiền hàng gồm VAT
        ws["G16"] = args.ship # Phí ship
        ws["G17"] = "=G15+G16" # Tổng cộng

        # Chi tiết vận chuyển trong phần điều khoản
        ws["C24"] = f"Giao COD về địa chỉ: {args.address}. Phí ship {args.ship:,.0f}đ."

        # Merge lại các ô theo toạ độ mới (không có cột E)
        new_merges = [
            "D1:G1", "D2:G2", "D3:G3", "D4:G4", # Tiêu đề
            "A6:C6", "D6:G6",                  # Khách hàng & Quy cách
            "A7:G7",                           # Dòng địa chỉ mới
            "E13:F13", "E14:F14", "E15:F15", "E16:F16", "E17:F17", # Khối tổng tiền nhãn
            "B18:G18",                         # Ghi chú header
            "C19:G19", "C20:G20", "C21:G21", "C22:G22", "C23:G23", "C24:G24", # Ghi chú values
            "B26:C26", "B28:C28",              # Chữ ký trái
            "F26:G26", "F28:G28"               # Chữ ký phải
        ]
        for m in new_merges:
            ws.merge_cells(m)

        # Cập nhật độ rộng các cột
        col_widths = {'A': 8, 'B': 40, 'C': 13, 'D': 12, 'E': 16, 'F': 20, 'G': 16}
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Thiết lập format hiển thị
        ws["D10"].number_format = "#,##0;(#,##0);'-'"
        ws["E10"].number_format = '#,##0.0" ₫"'
        ws["F10"].number_format = '#,##0" ₫"'
        ws["G10"].number_format = '#,##0.0" ₫"'
        ws["D11"].number_format = "#,##0;(#,##0);'-'"
        ws["F11"].number_format = '#,##0" ₫"'
        ws["G11"].number_format = '#,##0.0" ₫"'
        for r in range(13, 18):
            ws.cell(row=r, column=7).number_format = '#,##0" ₫"'
    else:
        print("Order is based on meters. Keeping Column E...")
        # Dòng Quy cách (dòng 6)
        ws["D6"] = f" Quy cách: Cuộn khổ 58cm | {args.meters:.2f} mét dài phân bổ bế bọc màng định hình"

        # Điền sản phẩm (Dòng 10)
        ws["B10"] = args.product
        ws["C10"] = args.size
        ws["D10"] = args.qty
        ws["E10"] = args.meters
        ws["F10"] = args.rate_excl_vat # Đơn giá mét trước VAT

        # Ghi đè lại công thức dòng 10
        ws["G10"] = "=E10*F10"
        ws["H10"] = "=G10/D10"

        # Ghi đè các công thức tính tổng dòng 11
        ws["D11"] = "=SUM(D10:D10)"
        ws["E11"] = "=SUM(E10:E10)"
        ws["G11"] = "=SUM(G10:G10)"
        ws["H11"] = "=G11/D11"

        ws["H13"] = "=G11"
        ws["H14"] = "=H13*0.08"
        ws["H15"] = "=H13+H14"
        ws["H16"] = args.ship
        ws["H17"] = "=H15+H16"

        # Chi tiết vận chuyển trong phần điều khoản
        ws["C24"] = f"Giao COD về địa chỉ: {args.address}. Phí ship {args.ship:,.0f}đ."

    # Lưu file Excel
    wb.save(excel_output)
    print(f"[OK] Excel saved at: {excel_output}")

    # 3. Xuất file PDF và TỰ ĐỘNG KIỂM TRA LỖI CÔNG THỨC (#VALUE!, #REF!)
    try:
        print("Opening Excel to calculate and convert to PDF...")
        excel_app = win32com.client.DispatchEx("Excel.Application")
        excel_app.Visible = False
        excel_app.DisplayAlerts = False
        
        abs_excel = os.path.abspath(excel_output)
        abs_pdf = os.path.abspath(pdf_output)
        
        wb_obj = excel_app.Workbooks.Open(abs_excel)
        ws_obj = wb_obj.ActiveSheet
        
        # Ép Excel tính toán lại tất cả công thức
        excel_app.CalculateFull()
        
        # --- KIỂM TRA LỖI XÁC THỰC ---
        print("Checking for formula errors (#VALUE!, #REF!, etc.)...")
        cells_to_check = (
            ["F10", "G10", "D11", "F11", "G11", "G13", "G14", "G15", "G16", "G17"]
            if not use_meters else
            ["G10", "H10", "D11", "E11", "G11", "H11", "H13", "H14", "H15", "H16", "H17"]
        )
        errors_found = []
        for cell in cells_to_check:
            val = ws_obj.Range(cell).Text
            if any(err in str(val) for err in ["#VALUE!", "#REF!", "#N/A", "#NAME?", "#DIV/0!", "#NUM!", "#NULL!"]):
                errors_found.append(f"Cell {cell} has error value: {val}")
        
        if errors_found:
            print("\n[ERR] DATABASE/EXCEL FORMULA ERROR DETECTED:")
            for err in errors_found:
                print(f"  - {err}")
            wb_obj.Close(False)
            excel_app.Quit()
            sys.exit(1)
            
        print("No formula errors found. Exporting to PDF...")
        
        ws_obj.PageSetup.Orientation = 1 # Portrait
        ws_obj.PageSetup.LeftMargin = excel_app.InchesToPoints(0.4)
        ws_obj.PageSetup.RightMargin = excel_app.InchesToPoints(0.4)
        ws_obj.PageSetup.TopMargin = excel_app.InchesToPoints(0.5)
        ws_obj.PageSetup.BottomMargin = excel_app.InchesToPoints(0.5)
        
        ws_obj.PageSetup.Zoom = False
        ws_obj.PageSetup.FitToPagesWide = 1
        ws_obj.PageSetup.FitToPagesTall = 1
        
        # Lưu PDF
        ws_obj.ExportAsFixedFormat(0, abs_pdf)
        print(f"[OK] PDF generated at: {pdf_output}")

        wb_obj.Close(False)
        excel_app.Quit()
    except Exception as e:
        print(f"[ERR] Execution failed: {e}")
        sys.exit(1)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Tạo báo giá Excel/PDF từ template chuẩn")
    parser.add_argument("--output_dir", required=True, help="Thư mục xuất báo giá")
    parser.add_argument("--name", required=True, help="Tên khách hàng")
    parser.add_argument("--phone", required=True, help="Số điện thoại")
    parser.add_argument("--address", required=True, help="Địa chỉ giao hàng")
    parser.add_argument("--product", required=True, help="Tên sản phẩm in tem")
    parser.add_argument("--size", required=True, help="Kích thước tem")
    parser.add_argument("--qty", type=int, required=True, help="Số lượng cái")
    parser.add_argument("--meters", type=float, default=None, help="Độ dài mét cuộn (nếu có)")
    parser.add_argument("--rate_excl_vat", type=float, required=True, help="Đơn giá in chưa VAT (theo mét hoặc theo cái/tờ)")
    parser.add_argument("--ship", type=float, required=True, help="Phí ship")
    parser.add_argument("--order_code", required=True, help="Mã đơn hàng")
    parser.add_argument("--spec_text", default=None, help="Văn bản quy cách thay thế")

    args = parser.parse_args()
    create_quote(args)
