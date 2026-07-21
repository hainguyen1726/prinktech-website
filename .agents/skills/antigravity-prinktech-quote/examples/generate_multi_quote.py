import openpyxl
import os
import shutil
import win32com.client
import sys
import argparse
import json
from openpyxl.styles import Font, Alignment, Border, Side

def create_multi_quote(args):
    template_file = r"D:\16. Code\32-website-prinktech\.agents\skills\antigravity-prinktech-quote\resources\template_bao_gia.xlsx"
    name_slug = args.name.replace(' ', '_')
    excel_output = os.path.join(args.output_dir, f"1. Bao_gia_in_tem_UV_DTF_{name_slug}_V2.xlsx")
    pdf_output = os.path.join(args.output_dir, f"2. Bao_gia_in_tem_UV_DTF_{name_slug}_V2.pdf")

    if not os.path.exists(template_file):
        print(f"[ERR] Template not found: {template_file}")
        sys.exit(1)

    os.makedirs(args.output_dir, exist_ok=True)
    
    # 1. Sao chép template Excel
    print("Copying template file...")
    shutil.copyfile(template_file, excel_output)

    # 2. Điền thông tin vào Excel
    print("Filling customer data...")
    wb = openpyxl.load_workbook(excel_output)
    ws = wb.active

    items = json.loads(args.items_json)
    num_items = len(items)

    # Cập nhật số đơn hàng ở D2
    ws["D2"] = f"Đơn hàng: {args.order_code}"
    # Khách hàng & Địa chỉ
    ws["A6"] = f" Khách hàng: {args.name}"
    ws["A7"] = f" Địa chỉ giao hàng: {args.address}"
    ws["D6"] = f" Quy cách: {args.spec_text or 'Tem UV DTF bọc màng định hình bế sẵn'}"

    # Unmerge toàn bộ sheet trước khi sửa đổi cấu trúc
    merged_ranges = list(ws.merged_cells.ranges)
    for r in merged_ranges:
        ws.unmerge_cells(r.coord)

    # Xoá cột E (Mét dài)
    ws.delete_cols(5, amount=1)

    # Nếu có nhiều hơn 1 sản phẩm, chèn thêm hàng vào sau dòng 10
    if num_items > 1:
        ws.insert_rows(11, amount=num_items - 1)

    # Border style cho các ô dữ liệu sản phẩm
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Điền từng sản phẩm
    for idx, item in enumerate(items):
        r = 10 + idx
        ws.cell(row=r, column=1, value=idx + 1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=2, value=item['product']).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=r, column=3, value=item['size']).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=4, value=item['qty']).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=r, column=5, value=item['rate']).alignment = Alignment(horizontal='right', vertical='center')
        
        ws.cell(row=r, column=6, value=f"=D{r}*E{r}").alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=r, column=7, value=f"=F{r}/D{r}").alignment = Alignment(horizontal='right', vertical='center')

        # Formatting
        ws.cell(row=r, column=4).number_format = "#,##0;(#,##0);'-'"
        ws.cell(row=r, column=5).number_format = '#,##0.0" ₫"'
        ws.cell(row=r, column=6).number_format = '#,##0" ₫"'
        ws.cell(row=r, column=7).number_format = '#,##0.0" ₫"'

        for c in range(1, 8):
            ws.cell(row=r, column=c).border = thin_border
            ws.cell(row=r, column=c).font = Font(name='Segoe UI', size=10)

    # Dòng Tổng cộng (nằm ở dòng 10 + num_items)
    tot_row = 10 + num_items
    ws.cell(row=tot_row, column=1, value="Tổng cộng").font = Font(name='Segoe UI', size=10, bold=True)
    ws.cell(row=tot_row, column=4, value=f"=SUM(D10:D{tot_row-1})").font = Font(name='Segoe UI', size=10, bold=True)
    ws.cell(row=tot_row, column=6, value=f"=SUM(F10:F{tot_row-1})").font = Font(name='Segoe UI', size=10, bold=True)
    ws.cell(row=tot_row, column=7, value=f"=F{tot_row}/D{tot_row}").font = Font(name='Segoe UI', size=10, bold=True)

    ws.cell(row=tot_row, column=4).number_format = "#,##0;(#,##0);'-'"
    ws.cell(row=tot_row, column=6).number_format = '#,##0" ₫"'
    ws.cell(row=tot_row, column=7).number_format = '#,##0.0" ₫"'

    # Các dòng tổng tiền VAT & Ship (Offset theo số hàng chèn)
    offset = num_items - 1
    r_sub = 13 + offset
    r_vat = 14 + offset
    r_incl = 15 + offset
    r_ship = 16 + offset
    r_tot = 17 + offset

    ws.cell(row=r_sub, column=6, value="Cộng tiền hàng (chưa VAT):")
    ws.cell(row=r_sub, column=7, value=f"=F{tot_row}")

    ws.cell(row=r_vat, column=6, value="Thuế VAT (8%):")
    ws.cell(row=r_vat, column=7, value=f"=G{r_sub}*0.08")

    ws.cell(row=r_incl, column=6, value="Tiền hàng (đã gồm 8% VAT):")
    ws.cell(row=r_incl, column=7, value=f"=G{r_sub}+G{r_vat}")

    ws.cell(row=r_ship, column=6, value="Phí vận chuyển:")
    ws.cell(row=r_ship, column=7, value=args.ship)

    ws.cell(row=r_tot, column=6, value="TỔNG THANH TOÁN (hàng + ship):")
    ws.cell(row=r_tot, column=7, value=f"=G{r_incl}+G{r_ship}")

    for r in range(r_sub, r_tot + 1):
        ws.cell(row=r, column=7).number_format = '#,##0" ₫"'

    # Chi tiết vận chuyển trong phần điều khoản
    r_term = 24 + offset
    ws.cell(row=r_term, column=3, value=f"Giao COD về địa chỉ: {args.address}. Phí ship {args.ship:,.0f}đ.")

    # Merge lại các ô theo toạ độ mới
    new_merges = [
        "D1:G1", "D2:G2", "D3:G3", "D4:G4", # Tiêu đề
        "A6:C6", "D6:G6",                  # Khách hàng & Quy cách
        "A7:G7",                           # Dòng địa chỉ mới
        f"A{tot_row}:C{tot_row}",          # Tổng cộng label
        f"E{r_sub}:F{r_sub}", f"E{r_vat}:F{r_vat}", f"E{r_incl}:F{r_incl}", f"E{r_ship}:F{r_ship}", f"E{r_tot}:F{r_tot}", # Khối tổng tiền nhãn
        f"B{18+offset}:G{18+offset}",     # Ghi chú header
        f"C{19+offset}:G{19+offset}", f"C{20+offset}:G{20+offset}", f"C{21+offset}:G{21+offset}", f"C{22+offset}:G{22+offset}", f"C{23+offset}:G{23+offset}", f"C{24+offset}:G{24+offset}", # Ghi chú values
        f"B{26+offset}:C{26+offset}", f"B{28+offset}:C{28+offset}", # Chữ ký trái
        f"F{26+offset}:G{26+offset}", f"F{28+offset}:G{28+offset}"  # Chữ ký phải
    ]
    for m in new_merges:
        ws.merge_cells(m)

    # Cập nhật độ rộng các cột
    col_widths = {'A': 8, 'B': 40, 'C': 15, 'D': 12, 'E': 16, 'F': 20, 'G': 16}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.views.sheetView[0].showGridLines = True

    # Lưu file Excel
    wb.save(excel_output)
    print(f"[OK] Excel saved at: {excel_output}")

    # 3. Xuất file PDF và TỰ ĐỘNG KIỂM TRA LỖI CÔNG THỰC
    try:
        print("Opening Excel to calculate and convert to PDF...")
        excel_app = win32com.client.DispatchEx("Excel.Application")
        excel_app.Visible = False
        excel_app.DisplayAlerts = False
        
        abs_excel = os.path.abspath(excel_output)
        abs_pdf = os.path.abspath(pdf_output)
        
        wb_obj = excel_app.Workbooks.Open(abs_excel)
        ws_obj = wb_obj.ActiveSheet
        
        # Ép Excel tính toán lại tất cả công thức
        excel_app.CalculateFull()
        
        # --- KIỂM TRA LỖI XÁC THỰC ---
        print("Checking for formula errors (#VALUE!, #REF!, etc.)...")
        cells_to_check = [
            f"F{tot_row}", f"G{tot_row}", f"D{tot_row}",
            f"G{r_sub}", f"G{r_vat}", f"G{r_incl}", f"G{r_ship}", f"G{r_tot}"
        ]
        errors_found = []
        for cell in cells_to_check:
            val = ws_obj.Range(cell).Text
            if any(err in str(val) for err in ["#VALUE!", "#REF!", "#N/A", "#NAME?", "#DIV/0!", "#NUM!", "#NULL!"]):
                errors_found.append(f"Cell {cell} has error value: {val}")
        
        if errors_found:
            print("\n[ERR] DATABASE/EXCEL FORMULA ERROR DETECTED:")
            for err in errors_found:
                print(f"  - {err}")
            wb_obj.Close(False)
            excel_app.Quit()
            sys.exit(1)
            
        print("No formula errors found. Exporting to PDF...")
        
        ws_obj.PageSetup.Orientation = 1 # Portrait
        ws_obj.PageSetup.LeftMargin = excel_app.InchesToPoints(0.4)
        ws_obj.PageSetup.RightMargin = excel_app.InchesToPoints(0.4)
        ws_obj.PageSetup.TopMargin = excel_app.InchesToPoints(0.5)
        ws_obj.PageSetup.BottomMargin = excel_app.InchesToPoints(0.5)
        
        ws_obj.PageSetup.Zoom = False
        ws_obj.PageSetup.FitToPagesWide = 1
        ws_obj.PageSetup.FitToPagesTall = 1
        
        # Lưu PDF
        ws_obj.ExportAsFixedFormat(0, abs_pdf)
        print(f"[OK] PDF generated at: {pdf_output}")

        wb_obj.Close(False)
        excel_app.Quit()
    except Exception as e:
        print(f"[ERR] Execution failed: {e}")
        sys.exit(1)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Tạo báo giá đa sản phẩm Excel/PDF từ template chuẩn")
    parser.add_argument("--output_dir", required=True, help="Thư mục xuất báo giá")
    parser.add_argument("--name", required=True, help="Tên khách hàng")
    parser.add_argument("--phone", required=True, help="Số điện thoại")
    parser.add_argument("--address", required=True, help="Địa chỉ giao hàng")
    parser.add_argument("--items_json", required=True, help="Danh sách sản phẩm dạng JSON string")
    parser.add_argument("--ship", type=float, required=True, help="Phí ship")
    parser.add_argument("--order_code", required=True, help="Mã đơn hàng")
    parser.add_argument("--spec_text", default=None, help="Văn bản quy cách thay thế")

    args = parser.parse_args()
    create_multi_quote(args)
